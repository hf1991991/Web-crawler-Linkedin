from scrapy import Request
from scrapy.spiders import CrawlSpider, Rule
from scrapy.spiders.init import InitSpider
import scrapy.http as Http
from scrapy.utils.spider import iterate_spider_output

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Side, Alignment, Border
from openpyxl.utils import get_column_letter

import os

import json

import colorama
from termcolor import cprint

SYSTEM_IS_WINDOWS = os.name != 'posix'

colorama.init()

if not SYSTEM_IS_WINDOWS:
    whiteprint = lambda x: cprint(x, 'white')
    warnprint = lambda x: cprint(x, 'yellow')
    checkprint = lambda x: cprint(x, 'green')
    errorprint = lambda x: cprint(x, 'red')
else:
    whiteprint = lambda x: cprint(x, 'white')
    warnprint = lambda x: whiteprint("🟡 %s" % x)
    checkprint = lambda x: whiteprint("✅ %s" % x)
    errorprint = lambda x: whiteprint("❌ %s" % x)

CELL_SIDE = Side(
    border_style="thin",
    color="000000"
)

CELL_BORDER = Border(
    top=CELL_SIDE,
    bottom=CELL_SIDE,
    right=CELL_SIDE,
    left=CELL_SIDE,
)

LEFT_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="left",
    wrap_text=True
)

CENTER_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="center",
    wrap_text=True
)

NORMAL_FONT_CELL = Font()

BIG_FONT_CELL = Font(
    size=18
)

LINKS_TABLE_STARTING_LINE = 10
USERS_TABLE_STARTING_LINE = 3

class LinkedinSpider(InitSpider):
    name = "linkedin"
    allowed_domains = ["linkedin.com"]

    workbook_filename = 'Links.xlsx'
    workbook = None

    only_crawl_new_links = None
    crawl_not_a_person = None

    user_name = None
    passwd = None
    user_line_on_excel = None
    possible_users = []
    users_whole_cycles = 0

    start_urls = []
    parsed_urls = []

    request_retries = {}

    unicode_dict = {}

    login_page = 'https://www.linkedin.com/uas/login'
        
    def init_request(self):
        # Obtém os dados do excel:
        self.workbook = load_workbook(filename=self.workbook_filename)
        # Arruma dados da tabela de usuários do excel:
        self.fix_users_sheet_data()
        # A partir dos dados do excel, associa valores às variaveis de login, assim como à dos links:
        if self.get_login_data_from_workbook() is not None: return
        if self.get_links_from_workbook() is not None: return
        # Aplica estilo no excel:
        self.apply_links_sheet_style()
        self.apply_users_sheet_style()
        # Lê os valores de conversão unicode:
        self.read_unicode_conversion()
        # Realiza o login:
        return self.attempt_login()

    def attempt_login(self):
        if self.cycle_possible_users() is not None: return
        return Request(url=self.login_page, callback=self.login, dont_filter=True)

    def fix_users_sheet_data(self):
        users_sheet = self.workbook['Usuários']
        line = USERS_TABLE_STARTING_LINE
        while users_sheet['B%i' % line].value is not None or users_sheet['C%i' % line].value:
            if users_sheet['D%i' % line].value is None:
                users_sheet['D%i' % line] = 0
            if (users_sheet['E%i' % line].value != 'Sim') and (users_sheet['E%i' % line].value != 'Não') and (users_sheet['E%i' % line].value != 'Não testado'):
                users_sheet['E%i' % line] = 'Não testado'
            if (users_sheet['F%i' % line].value != 'Sim') and (users_sheet['F%i' % line].value != 'Não') and (users_sheet['F%i' % line].value != 'Não testado'):
                users_sheet['F%i' % line] = 'Não testado'
            if users_sheet['G%i' % line].value is None:
                users_sheet['G%i' % line] = '---'
            line += 1
        self.workbook.save(self.workbook_filename)


    def get_login_data_from_workbook(self):
        # whiteprint('GET_LOGIN_DATA_FROM_WORKBOOK')

        def has_been_tested(item):
            return item['does_it_work'] == 'Sim'

        def times_used(item):
            return item['times_used']
        
        users_sheet = self.workbook['Usuários']
        self.possible_users = []
        line = USERS_TABLE_STARTING_LINE

        while True:
            login = {
                'email': users_sheet['B%i' % line].value,
                'password': users_sheet['C%i' % line].value,
                'times_used': users_sheet['D%i' % line].value,
                'does_it_work': users_sheet['E%i' % line].value,
                'line': line
            }
            if login['email'] is None or login['password'] is None:
                break
            if login['does_it_work'] != 'Não':
                self.possible_users.append(login)
            line += 1

        if len(self.possible_users) == 0: 
            print()
            errorprint('Não há mais usuários válidos.\nEntre na tabela do Excel para adicionar um usuário, ou arrumar algum que tenha gerado um erro.\n')
            return 'Zero'
        
        self.possible_users.sort(key=times_used)
        self.possible_users.sort(key=has_been_tested)

        return None

    def cycle_possible_users(self):
   
        if self.user_name is not None: whiteprint('Trocando de login...')  

        users_sheet = self.workbook['Usuários']

        if len(self.possible_users) == 0:
            self.users_whole_cycles += 1
            print()
            if self.users_whole_cycles > 3:
                errorprint('Todos os usuários válidos já foram testados 3 vezes.\nEntre na tabela do Excel para adicionar um usuário, ou arrumar algum que tenha gerado um erro.\n')
                return 'Não há mais usuários válidos para serem utilizados'
            else:
                warnprint('Não foi possível realizar login com nenhuma das contas funcionais.\nIniciando tentativa %i de 3.' % (self.users_whole_cycles + 1))
                if self.get_login_data_from_workbook() is not None: return 'Error'
            self.workbook.save(self.workbook_filename)  

        new_user = self.possible_users.pop(0)

        self.user_name = new_user['email']
        self.passwd = new_user['password']
        self.user_line_on_excel = new_user['line']

        whiteprint("\nLogin em uso:\n - Email: %s\n - Senha: %s\n" % (self.user_name, self.passwd))

        users_sheet['D%i' % self.user_line_on_excel] = new_user['times_used'] + 1
        self.workbook.save(self.workbook_filename)
        
        return None

    def get_links_from_workbook(self):
        # whiteprint('GET_LINKS_FROM_WORKBOOK')
        links_sheet = self.workbook['Links']

        self.only_crawl_new_links = links_sheet['D5'].value == 'Sim'
        self.crawl_not_a_person = links_sheet['D6'].value == 'Sim'

        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        while link is not None:
            if self.only_crawl_new_links:
                is_a_cell_empty = False
                for column in "BDEFGH":
                    if links_sheet['%s%i' % (column, line)].value == None:
                        is_a_cell_empty = True
                if is_a_cell_empty and ((links_sheet['B%i' % line].value != 'Não é uma pessoa') or self.crawl_not_a_person):
                    self.start_urls.append(link)
            else:
                if (links_sheet['B%i' % line].value != 'Não é uma pessoa') or self.crawl_not_a_person:
                    self.start_urls.append(link)
            line += 1
            link = links_sheet['C%i' % line].value
        if len(self.start_urls) == 0:
            print()
            checkprint('Todos os links do Excel já passaram pelo scraping!\nCaso queira recarregá-los, desative a configuração de "Apenas obter dados dos links cujos campos da linha estão vazios" e salve o arquivo\n')
            return 'Sem links para scraping'
        else:
            return None
        # whiteprint("start urls:\n")
        # whiteprint(self.start_urls)

    def apply_links_sheet_style(self):        
        self.apply_style_to_workbook_sheet(sheet=self.workbook['Links'], verification_column='C', starting_line=LINKS_TABLE_STARTING_LINE, columns="BCDEFGH")
        self.apply_style_to_workbook_sheet(sheet=self.workbook['Links'], alignment=CENTER_CELL_ALIGNMENT, font=BIG_FONT_CELL, verification_column='C', starting_line=LINKS_TABLE_STARTING_LINE, columns="B")

    def apply_users_sheet_style(self):        
        self.apply_style_to_workbook_sheet(sheet=self.workbook['Usuários'], verification_column='B', starting_line=USERS_TABLE_STARTING_LINE, columns="BCDEFG")

    def apply_style_to_workbook_sheet(self, sheet, verification_column, starting_line, columns, alignment=LEFT_CELL_ALIGNMENT, border=CELL_BORDER, font=NORMAL_FONT_CELL):
        # whiteprint('APPLY_STYLE_TO_WORKBOOK')
        line = starting_line
        link = sheet['%s%i' % (verification_column, line)].value
        while link is not None:
            for column in columns:
                cell = sheet['%s%i' % (column, line)]
                cell.alignment = alignment
                cell.border = border
                cell.font = font
            line += 1
            link = sheet['%s%i' % (verification_column, line)].value
        self.workbook.save(self.workbook_filename)

    def write_on_workbook(self, url, user_dict, page_exists):
        # whiteprint('WRITE_ON_WORKBOOK')
        links_sheet = self.workbook['Links']
        column_association = {
            'D': 'first_name',
            'E': 'last_name',
            'F': 'occupation',
            'G': 'location',
            'H': 'about',
        }
        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        while link is not None:
            if link == url:
                if user_dict is not None:
                    links_sheet['B%i' % line] = 'Sim'
                    for column in column_association:
                        text = user_dict[column_association[column]]
                        if text == None:
                            text = '---'
                        links_sheet['%s%i' % (column, line)] = text
                elif not page_exists:
                    links_sheet['B%i' % line] = 'Não é uma pessoa'
                else:
                    links_sheet['B%i' % line] = 'Não'
                self.workbook.save(self.workbook_filename)
                return
            line += 1
            link = links_sheet['C%i' % line].value
        whiteprint('write_on_workbook: foram obtidos os dados de %s, mas o link não foi encontrado na tabela.' % url)

    def read_unicode_conversion(self):
        with open('unicode_conversion.json', 'r') as f:
            self.unicode_dict = json.loads(f.read())

    def login(self, response):
        return Http.FormRequest.from_response(
            response,
            formdata={
                'session_key': self.user_name,
                'session_password': self.passwd,
            },
            callback = self.check_login_response,
            meta={
                'proxy': None
            }
        )

    def set_error_message_on_users_sheet(self, error_text, login_works, reusable_login):
        users_sheet = self.workbook['Usuários']

        if error_text is None: error_text = '---'

        users_sheet['G%i' % self.user_line_on_excel] = error_text

        if login_works:
            users_sheet['E%i' % self.user_line_on_excel] = 'Sim'
        else:
            users_sheet['E%i' % self.user_line_on_excel] = 'Não'

        if reusable_login:
            users_sheet['F%i' % self.user_line_on_excel] = 'Sim'
        else:
            users_sheet['F%i' % self.user_line_on_excel] = 'Não'
            
        self.workbook.save(self.workbook_filename)

    def response_is_ban(self, request, response):
        ban = False

        if "Your account has been restricted" in str(response.body):
            ban = False
        elif "Let&#39;s do a quick security check" in str(response.body):
            ban = True
        elif "The login attempt seems suspicious." in str(response.body):
            ban = True
        elif "that&#39;s not the right password" in str(response.body):
            ban = True
        elif "We’re unable to reach you" in str(response.body):
            ban = True
        elif '<meta name="isGuest" content="false" />' in str(response.body):
            ban = False
        else:
            ban = True

        return ban

    def exception_is_ban(self, request, exception):
        return None

    def check_login_response(self, response):
        logged_in = False
        error_text = None
        login_works = True
        reusable_login = True
        
        loginerrorprint = lambda x: warnprint('Login falhou. %s\n' % x)

        print()

        if "Your account has been restricted" in str(response.body):
            login_works = False
            reusable_login = False
            error_text = 'Conta bloqueada pelo Linkedin por muitas tentativas. Troque esta conta por outra, ou remova esta linha do Excel.'
            loginerrorprint('Conta bloqueada pelo Linkedin por muitas tentativas.\nPara mais detalhes, entre na aba "Usuários" do Excel.')
        elif "Let&#39;s do a quick security check" in str(response.body):
            login_works = False
            error_text = 'Conta pede uma verificação se é um robô. Troque esta conta por outra, ou remova esta linha do Excel.'
            loginerrorprint("Conta pede uma verificação de se é um robô")
        elif "The login attempt seems suspicious." in str(response.body):
            login_works = False
            error_text = 'Conta pede uma verificação se é um robô. Troque esta conta por outra, ou remova esta linha do Excel.'
            loginerrorprint("Conta pede que seja copiado um código do email")
        elif "that&#39;s not the right password" in str(response.body):
            login_works = False
            # save_to_file(
            #     "login.html",
            #     response.body
            # )
            error_text = 'A conta ou a senha parecem estar erradas. Verifique se o usuário e senha estão corretos.'
            loginerrorprint("A senha está errada.\nVerifique se o usuário e senha estão corretos.")
        elif "We’re unable to reach you" in str(response.body):
            login_works = False
            error_text = 'O Linkedin pediu uma verificação de email. Faça login com esta conta no browser e aperte "Skip".'
            loginerrorprint('O Linkedin pediu uma verificação de email.\nPara mais detalhes, entre na aba "Usuários" do Excel.')
        elif '<meta name="isGuest" content="false" />' in str(response.body):
            logged_in = True
            checkprint("Login realizado. Vamos começar o crawling!\n")
        else:
            login_works = False
            reusable_login = False
            error_text = 'Erro desconhecido.'
            loginerrorprint("Erro desconhecido.\n")

        self.set_error_message_on_users_sheet(error_text, login_works, reusable_login)

        if logged_in:
            return self.initialized() 
        else:
            return self.attempt_login()

    def start_requests(self):
        self._postinit_reqs = self.start_requests_without_proxy_change()
        return iterate_spider_output(self.init_request())

    def start_requests_without_proxy_change(self):
        # whiteprint('START_SPLASH_REQUESTS')
        for url in self.start_urls:
            # O seguinte código faz com que todos os Requests depois do login não mudem de proxy:
            yield Request(
                url=url, 
                callback=self.parse,
                meta={
                    'proxy': None
                }
            )

    def get_user_data_string(self, response):
        body = response.body.decode('utf8')
        user_data_string = '{&quot;birthDateOn' + str(body).split(',{&quot;birthDateOn')[-1]
        end = 1
        partial = user_data_string[:end]
        while (partial.count('{') != partial.count('}')) and (partial.count('{') < 200) and (len(user_data_string) > end):
            end += 1
            partial = user_data_string[:end]
            # if partial.endswith('{') or partial.endswith('}'):
            #     whiteprint(partial.count('{'), partial.count('}'))
        if partial.count('{') != partial.count('}'):
            whiteprint('ERRO em get_user_data_string: não foi possivel obter dados do usuário em %s' % response.url)
            return None
        return partial

    def parse(self, response):
        user_dict = None
        page_exists = False

        self.parsed_urls.append(response.url)

        counter = '(%i/%i) ' % (len(self.parsed_urls), len(self.start_urls))

        print()

        # Se página não tiver a seguinte string, ela provavelmente foi carregada errada:
        if 'www.linkedin.com/in/' not in str(response.url):
            errorprint('%sEste não é um link de um perfil: %s\n' % (counter, response.url))

        # Se página não tiver a seguinte string, ela provavelmente 
        # foi carregada errada, ou não é uma página válida:
        elif '{&quot;birthDateOn' not in str(response.body):
            retries = 0
            if str(response.url) in list(self.request_retries.keys()): retries = self.request_retries[str(response.url)]
            self.request_retries[str(response.url)] = retries + 1

            if retries < 1:
                self.start_urls.append(response.url)
                warnprint('%sErro no parsing de %s\nAdicionando novamente à fila de links...\n' % (counter, response.url))  
                return Request(url=response.url, callback=self.parse, dont_filter=True)
            else:
                errorprint('%sEste provavelmente não é um link de um perfil: %s' % (counter, response.url))  
                if not self.crawl_not_a_person: 
                    whiteprint('Caso seja, volte para o Excel e habilite a configuração "Tentar obter dados de páginas que foram marcadas como \'Não é uma pessoa\'\n')
            
        
        else:
            page_exists = True

            filename = 'profile-%s.json' % response.url.split("/")[4]

            user_data_string = self.get_user_data_string(response)

            if user_data_string != None:
                user_data = parse_text_to_json(user_data_string, self.unicode_dict, filename)

                if user_data is not None:
                    user_dict = {
                        'first_name': user_data['firstName'],
                        'last_name': user_data['lastName'],
                        'occupation': user_data['headline'],
                        'location': user_data['locationName'],
                        'about': user_data['summary'],
                    }
                    checkprint('%sParsing corretamente realizado em %s\n' % (counter, response.url))
                else:
                    errorprint('%sErro no parsing de %s\n' % (counter, response.url))

        self.write_on_workbook(response.url, user_dict, page_exists)


def parse_text_to_json(text, replacements, filename):
    try:
        text = convert_unicode(text, replacements)
        return json.loads(text)
    except Exception:
        # save_to_file(
        #     filename, 
        #     text
        # )
        return None

def convert_unicode(text, replacements):
    try:
        text = str(text)
        for unicode_char in list(replacements.keys()):
            for type in list(replacements[unicode_char].keys()):
                for element in replacements[unicode_char][type]:
                    text = text.replace(str(element), str(unicode_char))
    except Exception:
        print()
        errorprint('convert_unicode: não foi possível converter os caracteres unicode.\n')
    return text

def save_to_file(filename, element):
    element = str(element).replace("'", '"').replace('"s ', "'s ").replace('True', 'true').replace('False', 'false').replace('None', 'null')
    with open(filename, 'wb') as f:
        f.write(str.encode(str(element)))
    whiteprint('\n💽 Texto salvo como %s\n' % filename)
