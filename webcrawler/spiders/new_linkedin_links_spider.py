from scrapy import Request, Spider

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Side, Alignment, Border
from openpyxl.utils import get_column_letter

import json

whiteprint = lambda x: print("\033[97m{}\033[00m".format(x)) 
warnprint = lambda x: whiteprint("🟡 %s" % x)
checkprint = lambda x: whiteprint("✅ %s" % x)
errorprint = lambda x: whiteprint("❌ %s" % x)

CELL_SIDE = Side(
    border_style="thin",
    color="000000"
)

CELL_BORDER = Border(
    top=CELL_SIDE,
    bottom=CELL_SIDE,
    right=CELL_SIDE,
    left=CELL_SIDE,
)

LEFT_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="left",
    wrap_text=True
)

CENTER_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="center",
    wrap_text=True
)

NORMAL_FONT_CELL = Font()

BIG_FONT_CELL = Font(
    size=18
)

LINKS_TABLE_STARTING_LINE = 10
USERS_TABLE_STARTING_LINE = 3

class LinkedinSpider(Spider):
    name = "linkedin_links"

    workbook_filename = 'Links.xlsx'

    only_crawl_new_links = None

    user_name = None
    passwd = None
    user_line_on_excel = None
    possible_users = []
    users_whole_cycles = 0

    workbook = load_workbook(filename='Links.xlsx')

    start_urls = ['https://www.google.com/search?ei=z1byXp7KOLOo5OUPiOGJgAI&q=site%3Awww.linkedin.com%2Fin%2F&oq=site%3Awww.linkedin.com%2Fin%2F&gs_lcp=CgZwc3ktYWIQA1CHqAJY09kCYP7dAmgAcAB4AIAB7wGIAaAWkgEFMC41LjmYAQCgAQGqAQdnd3Mtd2l6&sclient=psy-ab&ved=0ahUKEwje7Zjp1JjqAhUzFLkGHYhwAiAQ4dUDCAw&uact=5']

    def write_on_workbook(self, new_links):
        links_sheet = self.workbook['Links']

        line = LINKS_TABLE_STARTING_LINE

        link = links_sheet['C%i' % line].value

        while (link is not None) or (len(new_links) > 0): 
            if link is None:
                links_sheet['C%i' % line] = new_links.pop()
            line += 1
            link = links_sheet['C%i' % line].value

        self.workbook.save(self.workbook_filename)

    def parse(self, response):
        
        links = []

        for link in response.css('div a::attr(href)').getall():
            if "linkedin.com" in str(link): 
                links.append(str(link))

        self.write_on_workbook(links)