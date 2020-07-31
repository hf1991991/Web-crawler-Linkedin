from scrapy import Request
from scrapy.spiders import CrawlSpider, Rule
from scrapy.spiders.init import InitSpider
import scrapy.http as Http
from scrapy.utils.spider import iterate_spider_output

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Side, Alignment, Border
from openpyxl.utils import get_column_letter

import colorama
from termcolor import cprint

from ..unicode_conversion import unicode_dict

import os

import json

SYSTEM_IS_WINDOWS = os.name != 'posix'

colorama.init()

if SYSTEM_IS_WINDOWS:
    def whiteprint(x): return cprint(x, 'magenta')
    def warnprint(x): return cprint('Aviso: %s' % x, 'yellow')
    def checkprint(x): return cprint(x, 'green')
    def errorprint(x): return cprint('Erro: %s' % x, 'red')
else:
    def whiteprint(x): return cprint(x, 'white')
    def warnprint(x): return whiteprint("🟡 %s" % x)
    def checkprint(x): return whiteprint("✅ %s" % x)
    def errorprint(x): return whiteprint("❌ %s" % x)

CELL_SIDE = Side(
    border_style="thin",
    color="000000"
)

CELL_BORDER = Border(
    top=CELL_SIDE,
    bottom=CELL_SIDE,
    right=CELL_SIDE,
    left=CELL_SIDE,
)

LEFT_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="left",
    wrap_text=True
)

CENTER_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="center",
    wrap_text=True
)

NORMAL_FONT_CELL = Font()

BIG_FONT_CELL = Font(
    size=18
)

LINKS_TABLE_STARTING_LINE = 10
USERS_TABLE_STARTING_LINE = 3


class LinkedinSpider(InitSpider):
    name = "linkedin"
    handle_httpstatus_list = [999]

    workbook_filename = 'Links.xlsx'
    workbook = None

    only_crawl_new_links = None
    crawl_not_a_person = None

    user_name = None
    password = None
    user_line_on_excel = None
    possible_users = []

    start_urls = []
    parsed_urls = []

    request_retries = {}

    company_parsed_data = {
        'nome': None,
        'funcionarios': [],
    }
    parsed_data = {
        'empresas': []
    }

    login_page = 'https://www.linkedin.com/uas/login'

    def __init__(self, excel_file):
        self.workbook_filename = excel_file

    def init_request(self):
        # save_to_file(
        #     'following_json.json',
        #     convert_unicode(convert_txt, unicode_dict)
        # )
        # Obtém os dados do excel:
        self.read_excel()
        # Arruma dados da tabela de usuários do excel:
        self.fix_users_sheet_data()
        # Arruma os links que não começam com www:
        self.fix_links_without_www()
        # A partir dos dados do excel, associa valores às variaveis de login, assim como à dos links:
        if self.get_login_data_from_workbook() is not None:
            return
        if self.get_links_from_workbook() is not None:
            return
        # Verifica se há e remove links duplicados:
        self.check_for_duplicate_links()
        # Aplica estilo no excel:
        self.apply_links_sheet_style()
        self.apply_users_sheet_style()
        # Realiza o login:
        return self.attempt_login()

    def read_excel(self):
        self.workbook = load_workbook(filename=self.workbook_filename)

    def attempt_login(self):
        if self.cycle_possible_users() is not None:
            return
        return Request(url=self.login_page, callback=self.login, dont_filter=True)

    def fix_users_sheet_data(self):
        users_sheet = self.workbook['Usuários']
        line = USERS_TABLE_STARTING_LINE
        while users_sheet['B%i' % line].value is not None or users_sheet['C%i' % line].value:
            if users_sheet['D%i' % line].value is None:
                users_sheet['D%i' % line] = 0
            if (users_sheet['E%i' % line].value != 'Sim') and (users_sheet['E%i' % line].value != 'Não') and (users_sheet['E%i' % line].value != 'Não testado'):
                users_sheet['E%i' % line] = 'Não testado'
            if (users_sheet['F%i' % line].value != 'Sim') and (users_sheet['F%i' % line].value != 'Não') and (users_sheet['F%i' % line].value != 'Não testado'):
                users_sheet['F%i' % line] = 'Não testado'
            if users_sheet['G%i' % line].value is None:
                users_sheet['G%i' % line] = '---'
            line += 1
        self.workbook.save(self.workbook_filename)

    def fix_links_without_www(self):
        links_sheet = self.workbook['Links']
        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        while link is not None:
            if ('www.linkedin' not in link) and ('linkedin' in link):
                novo_link = 'https://www.linkedin' + link.split('linkedin')[1]
                print()
                warnprint(
                    'O seguinte link não contém "www.linkedin.com": %s\nModificando-o para: %s' % (link, novo_link))
                links_sheet['C%i' % line] = novo_link
            line += 1
            link = links_sheet['C%i' % line].value
        self.workbook.save(self.workbook_filename)

    def check_for_duplicate_links(self):
        for link in self.start_urls:
            while self.start_urls.count(link) > 1:
                self.start_urls.remove(link)
                print()
                warnprint('Há uma cópia de link: %s' % link)

    def get_login_data_from_workbook(self):
        # whiteprint('GET_LOGIN_DATA_FROM_WORKBOOK')

        def has_been_tested(item):
            return item['does_it_work'] == 'Sim'

        def times_used(item):
            return item['times_used']

        users_sheet = self.workbook['Usuários']
        self.possible_users = []
        line = USERS_TABLE_STARTING_LINE

        while True:
            login = {
                'email': users_sheet['B%i' % line].value,
                'password': users_sheet['C%i' % line].value,
                'times_used': users_sheet['D%i' % line].value,
                'does_it_work': users_sheet['E%i' % line].value,
                'line': line
            }
            if login['email'] is None or login['password'] is None:
                break
            if login['does_it_work'] != 'Não':
                self.possible_users.append(login)
            line += 1

        if len(self.possible_users) == 0:
            print()
            errorprint(
                'Não há mais usuários válidos.\nEntre na tabela do Excel para adicionar um usuário, ou arrumar algum que tenha gerado um erro.\n')
            return 'Zero'

        self.possible_users.sort(key=times_used)
        self.possible_users.sort(key=has_been_tested)

        return None

    def cycle_possible_users(self):

        if self.user_name is not None:
            whiteprint('Trocando de login...')

        users_sheet = self.workbook['Usuários']

        if len(self.possible_users) == 0:
            print()
            errorprint('Todos os usuários válidos já foram testados.\nEntre na tabela do Excel para adicionar um usuário, ou arrumar algum que tenha gerado um erro.\n')
            self.workbook.save(self.workbook_filename)
            return 'Não há mais usuários válidos para serem utilizados'

        new_user = self.possible_users.pop(0)

        self.user_name = new_user['email']
        self.password = new_user['password']
        self.user_line_on_excel = new_user['line']

        whiteprint("\nLogin em uso:\n - Email: %s\n - Senha: %s\n" %
                   (self.user_name, self.password))

        users_sheet['D%i' %
                    self.user_line_on_excel] = new_user['times_used'] + 1
        self.workbook.save(self.workbook_filename)

        return None

    def get_links_from_workbook(self):
        # whiteprint('GET_LINKS_FROM_WORKBOOK')
        links_sheet = self.workbook['Links']

        self.only_crawl_new_links = links_sheet['D5'].value == 'Sim'
        self.crawl_not_a_person = links_sheet['D6'].value == 'Sim'

        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        while link is not None:
            link_data_state = None if links_sheet['B%i' % line].value is None else links_sheet['B%i' % line].value.replace(
                ' (Cópia)', '')
            if self.only_crawl_new_links:
                is_a_cell_empty = False
                for column in "BDEFGH":
                    if links_sheet['%s%i' % (column, line)].value == None:
                        is_a_cell_empty = True
                if is_a_cell_empty and ((link_data_state != 'Não é uma pessoa') or self.crawl_not_a_person):
                    self.start_urls.append(link)
            else:
                if (link_data_state != 'Não é uma pessoa') or self.crawl_not_a_person:
                    self.start_urls.append(link)
            line += 1
            link = links_sheet['C%i' % line].value

        if len(self.start_urls) == 0:
            print()
            checkprint('Todos os links do Excel já passaram pelo scraping!\nCaso queira recarregá-los, desative a configuração de "Apenas obter dados dos links cujos campos da linha estão vazios" e salve o arquivo\n')
            return 'Sem links para scraping'
        else:
            return None
        # whiteprint("start urls:\n")
        # whiteprint(self.start_urls)

    def apply_links_sheet_style(self):
        self.apply_style_to_workbook_sheet(
            sheet=self.workbook['Links'], verification_column='C', starting_line=LINKS_TABLE_STARTING_LINE, columns="BCDEFGH")
        self.apply_style_to_workbook_sheet(sheet=self.workbook['Links'], alignment=CENTER_CELL_ALIGNMENT,
                                           font=BIG_FONT_CELL, verification_column='C', starting_line=LINKS_TABLE_STARTING_LINE, columns="B")

    def apply_users_sheet_style(self):
        self.apply_style_to_workbook_sheet(
            sheet=self.workbook['Usuários'], verification_column='B', starting_line=USERS_TABLE_STARTING_LINE, columns="BCDEFG")

    def apply_style_to_workbook_sheet(self, sheet, verification_column, starting_line, columns, alignment=LEFT_CELL_ALIGNMENT, border=CELL_BORDER, font=NORMAL_FONT_CELL):
        # whiteprint('APPLY_STYLE_TO_WORKBOOK')
        line = starting_line
        link = sheet['%s%i' % (verification_column, line)].value
        while link is not None:
            for column in columns:
                cell = sheet['%s%i' % (column, line)]
                cell.alignment = alignment
                cell.border = border
                cell.font = font
            line += 1
            link = sheet['%s%i' % (verification_column, line)].value
        self.workbook.save(self.workbook_filename)

    def write_on_workbook(self, url, user_dict, page_exists):
        # whiteprint('WRITE_ON_WORKBOOK')
        links_sheet = self.workbook['Links']

        column_association = {
            'D': 'first_name',
            'E': 'last_name',
            'F': 'occupation',
            'G': 'location',
            'H': 'about',
        }

        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value

        link_count = 0

        while link is not None:
            if link == url:
                if user_dict is not None:
                    links_sheet['B%i' % line] = 'Sim'
                    for column in column_association:
                        text = user_dict[column_association[column]]
                        if text == None:
                            text = '---'
                        links_sheet['%s%i' % (column, line)] = text
                elif not page_exists:
                    links_sheet['B%i' % line] = 'Não é uma pessoa'
                else:
                    links_sheet['B%i' % line] = 'Não'
                link_count += 1
                if link_count > 1:
                    links_sheet['B%i' % line] = links_sheet['B%i' %
                                                            line].value + ' (Cópia)'
            line += 1
            link = links_sheet['C%i' % line].value

        self.workbook.save(self.workbook_filename)

        if link_count == 0:
            whiteprint(
                'write_on_workbook: foram obtidos os dados de %s, mas o link não foi encontrado na tabela.' % url)

    def login(self, response):
        return Http.FormRequest.from_response(
            response,
            formdata={
                'session_key': self.user_name,
                'session_password': self.password,
            },
            callback=self.check_login_response,
            meta={
                'proxy': None
            }
        )

    def set_error_message_on_users_sheet(self, error_text, login_works, reusable_login):
        users_sheet = self.workbook['Usuários']

        if error_text is None:
            error_text = '---'

        users_sheet['G%i' % self.user_line_on_excel] = error_text

        if login_works:
            users_sheet['E%i' % self.user_line_on_excel] = 'Sim'
        else:
            users_sheet['E%i' % self.user_line_on_excel] = 'Não'

        if reusable_login:
            users_sheet['F%i' % self.user_line_on_excel] = 'Sim'
        else:
            users_sheet['F%i' % self.user_line_on_excel] = 'Não'

        self.workbook.save(self.workbook_filename)

    # def response_is_ban(self, request, response):
    #     ban = False

    #     if "Your account has been restricted" in str(response.body):
    #         ban = False
    #     elif "Let&#39;s do a quick security check" in str(response.body):
    #         ban = True
    #     elif "The login attempt seems suspicious." in str(response.body):
    #         ban = True
    #     elif "that&#39;s not the right password" in str(response.body):
    #         ban = True
    #     elif "We’re unable to reach you" in str(response.body):
    #         ban = True
    #     elif '<meta name="isGuest" content="false" />' in str(response.body):
    #         ban = False
    #     else:
    #         ban = True

    #     return ban

    # def exception_is_ban(self, request, exception):
    #     return None

    def check_login_response(self, response):
        logged_in = False
        error_text = None
        login_works = True
        reusable_login = True

        def loginerrorprint(x): return warnprint('Login falhou. %s%s\n' % (
            x, '\nPara mais detalhes, entre na aba "Usuários" do Excel.' if not login_works else ''))

        print()

        if "Your account has been restricted" in str(response.body):
            login_works = False
            reusable_login = False
            error_text = 'Conta bloqueada pelo Linkedin por muitas tentativas. Troque esta conta por outra, ou remova esta linha do Excel.'
            loginerrorprint(
                'Conta bloqueada pelo Linkedin por muitas tentativas.')
        elif "Let&#39;s do a quick security check" in str(response.body):
            login_works = False
            error_text = 'Conta pede uma verificação se é um robô. Acesse o linkedin com essa conta e resolva o captcha.'
            loginerrorprint("Conta pede uma verificação de se é um robô")
        elif "The login attempt seems suspicious." in str(response.body):
            login_works = False
            error_text = 'Conta pede que seja copiado um código do email. Acesse o linkedin com essa conta e resolva o captcha.'
            loginerrorprint("Conta pede que seja copiado um código do email")
        elif "that&#39;s not the right password" in str(response.body):
            login_works = False
            # save_to_file(
            #     "login.html",
            #     response.body
            # )
            error_text = 'A conta ou a senha parecem estar erradas. Verifique se o usuário e senha estão corretos.'
            loginerrorprint(
                "A senha está errada.\nVerifique se o usuário e senha estão corretos.")
        elif "We’re unable to reach you" in str(response.body):
            login_works = False
            error_text = 'O Linkedin pediu uma verificação de email. Faça login com esta conta no browser e aperte "Skip".'
            loginerrorprint('O Linkedin pediu uma verificação de email.')
        # elif '<meta name="isGuest" content="false" />' in str(response.body):
        #     logged_in = True
        #     checkprint("Login realizado. Vamos começar o crawling!\n")
        else:
            logged_in = True
            checkprint("Login realizado. Vamos começar o crawling!\n")

        self.set_error_message_on_users_sheet(
            error_text, login_works, reusable_login)

        if logged_in:
            return self.initialized()
        else:
            return self.attempt_login()

    def start_requests(self):
        self._postinit_reqs = self.start_requests_without_proxy_change()
        return iterate_spider_output(self.init_request())

    def start_requests_without_proxy_change(self):
        # whiteprint('START_SPLASH_REQUESTS')
        for url in self.start_urls:
            # O seguinte código faz com que todos os Requests depois do login não mudem de proxy:
            yield Request(
                url=url,
                callback=self.parse,
                meta={
                    'proxy': None
                }
            )

    def get_big_json_included_array(self, response):
        body = str(response.body.decode('utf8'))

        birthIndex = body.rindex(',{&quot;birthDateOn')
        start = body[:birthIndex].rindex('<code ')
        end = body[start:].index('</code>') + start

        while (not body[start:end].startswith('{')) and start < end:
            start += 1

        while (not body[start:end].endswith('}')) and start < end:
            end -= 1

        if start >= end:
            whiteprint(
                'ERRO em get_big_json_included_array: não foi possivel obter dados do usuário em %s' % response.url)
            return None

        # save_to_file(
        #     response.url.split('/')[4] + '.html',
        #     convert_unicode(body[start:end], unicode_dict)
        # )

        return parse_text_to_json(body[start:end], unicode_dict, 'aa.json')["included"]

    def get_following_json_dictionary(self, response):
        body = str(response.body.decode('utf8'))

        birthIndex = body.rindex('&quot;followersCount&quot;:')
        start = body[:birthIndex].rindex('<code ')
        end = body[start:].index('</code>') + start

        while (not body[start:end].startswith('{')) and start < end:
            start += 1

        while (not body[start:end].endswith('}')) and start < end:
            end -= 1

        if start >= end:
            whiteprint(
                'ERRO em get_following_json_dictionary: não foi possivel obter dados do usuário em %s' % response.url)
            return None

        # save_to_file(
        #     response.url.split('/')[4] + '.html',
        #     convert_unicode(body[start:end], unicode_dict)
        # )

        return parse_text_to_json(body[start:end], unicode_dict, 'aa.json')

    def get_object_by_type(self, included_array, obj_type):
        array = []
        for obj in included_array:
            if obj['$type'] == obj_type:
                array.append(obj)
        return array

    def convert_date(self, date):
        return {
                'mes': date['month'] if 'month' in date else None,
                'ano': date['year'] if 'year' in date else None
            } if date is not None else None,

    def convert_date_range(self, date_range):
        return {
            'inicio': self.convert_date((date_range['start']) if ('start' in date_range) else None),
            'fim': self.convert_date((date_range['end']) if ('end' in date_range) else None)
        } if date_range is not None else None

    def format_conections(self, connections):
        return {
            'numero_exato': connections if connections != 500 else None,
            'minimo': connections if connections == 500 else None
        }

    def parse(self, response):
        user_dict = None
        page_exists = False

        self.parsed_urls.append(response.url)

        counter = '(%i/%i) ' % (len(self.parsed_urls), len(self.start_urls))

        print()

        # Se página não tiver a seguinte string, ela provavelmente foi carregada errada:
        if 'linkedin.com/in/' not in str(response.url):
            errorprint('%sEste não é um link de um perfil: %s\n' %
                       (counter, response.url))

        # Se página não tiver a seguinte string, ela provavelmente
        # foi carregada errada, ou não é uma página válida:
        elif '{&quot;birthDateOn' not in str(response.body):
            retries = 0
            if str(response.url) in list(self.request_retries.keys()):
                retries = self.request_retries[str(response.url)]
            self.request_retries[str(response.url)] = retries + 1

            if retries < 1:
                self.start_urls.append(response.url)
                warnprint('%sErro no parsing de %s\nAdicionando novamente à fila de links...\n' % (
                    counter, response.url))
                return Request(url=response.url, callback=self.parse, dont_filter=True)
            else:
                errorprint('%sEste provavelmente não é um link de um perfil: %s' % (
                    counter, response.url))
                if not self.crawl_not_a_person:
                    whiteprint(
                        'Caso seja, volte para o Excel e habilite a configuração "Tentar obter dados de páginas que foram marcadas como \'Não é uma pessoa\'\n')

        else:

            # try:

            page_exists = True

            # save_to_file(
            #     response.url.split('/')[4] + '.html',
            #     response.body
            # )

            included_array = self.get_big_json_included_array(response)

            if included_array is None:
                raise Exception('Erro com included_array')

            following_json = self.get_following_json_dictionary(response)

            if following_json is None:
                raise Exception('Erro com following_json')

            user_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Profile')[0]
            education_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Education')
            positions_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Position')
            volunteer_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.VolunteerExperience')
            skills_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Skill')
            honors_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Honor')
            projects_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Project')
            courses_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Course')
            languages_data = self.get_object_by_type(
                included_array, 'com.linkedin.voyager.dash.identity.profile.Language')

            # Itens que não estão em big_json:
            # com.linkedin.voyager.common.FollowingInfo

            user_dict = {
                'nome': user_data['firstName'] if 'firstName' in user_data else None,
                'sobrenome': user_data['lastName'] if 'lastName' in user_data else None,
                'cargo_atual': user_data['headline'] if 'headline' in user_data else None,
                'localizacao_atual': user_data['locationName'] if 'locationName' in user_data else None,
                'sobre': user_data['summary'] if 'summary' in user_data else None,
                'seguidores': following_json['data']['followersCount'],
                'conexoes': self.format_conections(following_json['data']['connectionsCount']),
                'habilidades': [skill['name'] for skill in skills_data],
                'linguas': [language['name'] for language in languages_data],
                'cursos_feitos': [course['name'] for course in courses_data],
                'premios': [
                    {
                        'nome': honor['title'] if 'title' in honor else None,
                        'instituicao': honor['issuer'] if 'issuer' in honor else None,
                        'descricao': honor['description'] if 'description' in honor else None,
                        'data': self.convert_date(
                            honor['issuedOn'] if 'issuedOn' in honor else None,
                        )
                    } for honor in honors_data
                ],
                'estudos': [
                    {
                        'instituicao': experience['schoolName'] if 'schoolName' in experience else None,
                        'formacao': experience['fieldOfStudy'] if 'fieldOfStudy' in experience else None,
                        'tilulo_obtido': experience['degreeName'] if 'degreeName' in experience else None,
                        'descricao': experience['description'] if 'description' in experience else None,
                        'periodo': self.convert_date_range(
                            experience['dateRange'] if 'dateRange' in experience else None
                        ),
                    } for experience in education_data
                ],
                'experiencia_profissional': [
                    {
                        'instituicao': experience['companyName'] if 'companyName' in experience else None,
                        'cargo': experience['title'] if 'title' in experience else None,
                        'descricao': experience['description'] if 'description' in experience else None,
                        'periodo': self.convert_date_range(
                            experience['dateRange'] if 'dateRange' in experience else None
                        ),
                    } for experience in positions_data
                ],
                'voluntariado': [
                    {
                        'instituicao': experience['companyName'] if 'companyName' in experience else None,
                        'papel': experience['role'] if 'role' in experience else None,
                        'causa': experience['cause'] if 'cause' in experience else None,
                        'descricao': experience['description'] if 'description' in experience else None,
                        'periodo': self.convert_date_range(
                            experience['dateRange'] if 'dateRange' in experience else None
                        ),
                    } for experience in volunteer_data
                ],
                'projetos': [
                    {
                        'titulo': project['title'],
                        'url': project['url'],
                        'descricao': project['description'],
                        'periodo': self.convert_date_range(
                            project['dateRange']
                        )
                    } for project in projects_data
                ],
                'dados_obtidos': True
            }

            checkprint('%sParsing corretamente realizado em %s\n' %
                    (counter, response.url))

            # except Exception:

                # user_dict = {
                #     'dados_obtidos': False
                # }

                # errorprint('%sErro no parsing de %s\n' % (counter, response.url))

        # print(user_dict)

        # for company in self.company_parsed_data:
        #     if company['name'] == 

        self.company_parsed_data['funcionarios'].append(user_dict)

        save_to_file(
            'company_parsed_data.json',
            json.dumps(self.company_parsed_data),
        )


def parse_text_to_json(text, replacements, filename):
    try:
        text = convert_unicode(text, replacements)
        return json.loads(text)
    except Exception:
        # save_to_file(
        #     filename,
        #     text
        # )
        return None


def convert_unicode(text, replacements):
    try:
        text = str(text)
        for unicode_char in list(replacements.keys()):
            for type in list(replacements[unicode_char].keys()):
                for element in replacements[unicode_char][type]:
                    text = text.replace(str(element), str(unicode_char))
    except Exception:
        print()
        errorprint(
            'convert_unicode: não foi possível converter os caracteres unicode.\n')
    return text


def save_to_file(filename, element):
    element = str(element).replace("'", '"').replace('"s ', "'s ").replace(
        'True', 'true').replace('False', 'false').replace('None', 'null')
    with open(filename, 'wb') as f:
        f.write(str.encode(str(element)))
    whiteprint('\n💽 Texto salvo como %s\n' % filename)


# def cmp_to_key(mycmp):
#     'Convert a cmp= function into a key= function'
#     class K:
#         def __init__(self, obj, *args):
#             self.obj = obj

#         def __lt__(self, other):
#             return mycmp(self.obj, other.obj) < 0

#         def __gt__(self, other):
#             return mycmp(self.obj, other.obj) > 0

#         def __eq__(self, other):
#             return mycmp(self.obj, other.obj) == 0

#         def __le__(self, other):
#             return mycmp(self.obj, other.obj) <= 0

#         def __ge__(self, other):
#             return mycmp(self.obj, other.obj) >= 0

#         def __ne__(self, other):
#             return mycmp(self.obj, other.obj) != 0
#     return K


def get_next_column(column):
    if len(column) == 0:
        return 'A'
    return (column[:-1] + chr(ord(column[-1]) + 1) if column[-1] != 'Z' else '%sA' % get_next_column(column[:-1]))

convert_txt = '{&quot;data&quot;:{&quot;distance&quot;:{&quot;value&quot;:&quot;OUT_OF_NETWORK&quot;,&quot;$type&quot;:&quot;com.linkedin.voyager.common.MemberDistance&quot;},&quot;entityUrn&quot;:&quot;urn:li:fs_profileNetworkInfo:ACoAAAWEek4BajqDsIutopY9XM4GurI8O2ewi7Q&quot;,&quot;following&quot;:false,&quot;followable&quot;:true,&quot;*followingInfo&quot;:&quot;urn:li:fs_followingInfo:urn:li:member:ACoAAAWEek4BajqDsIutopY9XM4GurI8O2ewi7Q&quot;,&quot;followersCount&quot;:9941,&quot;connectionsCount&quot;:500,&quot;$type&quot;:&quot;com.linkedin.voyager.identity.profile.ProfileNetworkInfo&quot;},&quot;included&quot;:[{&quot;entityUrn&quot;:&quot;urn:li:fs_followingInfo:urn:li:member:ACoAAAWEek4BajqDsIutopY9XM4GurI8O2ewi7Q&quot;,&quot;following&quot;:false,&quot;trackingUrn&quot;:&quot;urn:li:member:92568142&quot;,&quot;followerCount&quot;:9941,&quot;followingCount&quot;:null,&quot;$type&quot;:&quot;com.linkedin.voyager.common.FollowingInfo&quot;}]}'