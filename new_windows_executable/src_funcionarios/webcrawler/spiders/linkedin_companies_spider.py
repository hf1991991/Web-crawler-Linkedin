from scrapy import Request
from scrapy.exceptions import CloseSpider
from scrapy.spiders import CrawlSpider, Rule
from scrapy.spiders.init import InitSpider
import scrapy.http as Http
from scrapy.utils.spider import iterate_spider_output

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Side, Alignment, Border
from openpyxl.utils import get_column_letter

import colorama
from termcolor import cprint

from ..unicode_conversion import unicode_dict

import os

import json
from math import ceil

SYSTEM_IS_WINDOWS = os.name != 'posix'

colorama.init()

if SYSTEM_IS_WINDOWS:
    def whiteprint(x): return cprint(x, 'magenta')
    def warnprint(x): return cprint('Aviso: %s' % x, 'yellow')
    def checkprint(x): return cprint(x, 'green')
    def errorprint(x): return cprint('Erro: %s' % x, 'red')
else:
    def whiteprint(x): return cprint(x, 'white')
    def warnprint(x): return whiteprint("🟡 %s" % x)
    def checkprint(x): return whiteprint("✅ %s" % x)
    def errorprint(x): return whiteprint("❌ %s" % x)

CELL_SIDE = Side(
    border_style="thin",
    color="000000"
)

CELL_BORDER = Border(
    top=CELL_SIDE,
    bottom=CELL_SIDE,
    right=CELL_SIDE,
    left=CELL_SIDE,
)

LEFT_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="left",
    wrap_text=True
)

CENTER_CELL_ALIGNMENT = Alignment(
    vertical="center",
    horizontal="center",
    wrap_text=True
)

NORMAL_FONT_CELL = Font()

BIG_FONT_CELL = Font(
    size=18
)

LINKS_TABLE_STARTING_LINE = 8

class LinkedinSpider(InitSpider):
    name = "linkedin_companies"
    handle_httpstatus_list = [999]

    companies_excel_path = None
    workbook = None
    
    user_line_on_excel = None
    possible_users = []

    company_urls = []

    request_retries = {}

    login_page = 'https://www.linkedin.com/uas/login'

    def __init__(self, username, password, cookies_path, companies_excel_path='../macos_executable/Empresas.xlsx', employees_json_path=None):
        self.username = username
        self.password = password

        self.parse_cookies(cookies_path)

        self.companies_excel_path = companies_excel_path
        self.employees_json_path = employees_json_path

        self.employees_json_data = read_json_file(employees_json_path)

        if self.employees_json_data is None:
            self.employees_json_data = { 'empresas': [] }
        
        self.stored_requests = []
        self.companies_parsed = 0

    def init_request(self):
        # Obtém os dados do excel:
        self.read_excel()
        # Arruma os links que não começam com www:
        self.fix_links_without_www()
        # A partir dos dados do excel, carrega os links:
        if self.get_links_from_workbook() is not None:
            return
        # Verifica se há e remove links duplicados:
        self.check_for_duplicate_links()
        # Aplica estilo no excel:
        self.apply_links_sheet_style()
        # Realiza o login:
        return self.attempt_login()

    def parse_cookies(self, cookies_path):
        cookies = read_json_file(cookies_path)
        self.chrome_cookies = {}
        if cookies is None: return
        for line in cookies.split('\n'):
            if not line.startswith('#'):
                self.chrome_cookies[line.split()[-2]] = line.split()[-1]

    # Talvez seja interessante implementar headers também
    def cookie_request(self, url, priority=0, callback=None, cookies=None, meta=None, dont_filter=False):
        return Request(
            url=url,
            priority=priority,
            callback=callback,
            dont_filter=dont_filter,
            meta=meta,
            cookies=self.chrome_cookies
        )

    def read_excel(self):
        self.workbook = load_workbook(filename=self.companies_excel_path)

    def attempt_login(self):
        return self.cookie_request(url=self.login_page, callback=self.login, dont_filter=True)

    def fix_links_without_www(self):
        links_sheet = self.workbook['Empresas']
        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        while link is not None:
            if ('www.linkedin' not in link) and ('linkedin' in link):
                novo_link = 'https://www.linkedin' + link.split('linkedin')[1]
                print()
                warnprint(
                    'O seguinte link não contém "www.linkedin.com": %s\nModificando-o para: %s' % (link, novo_link))
                links_sheet['C%i' % line] = novo_link
            line += 1
            link = links_sheet['C%i' % line].value
        self.workbook.save(self.companies_excel_path)

    def count_employees_with_url(self, company_id):
        count = 0
        for company in self.employees_json_data['empresas']:
            if company['company_id'] == company_id:
                for employee in company['funcionarios']:
                    if ('url' in employee) and employee['url'] is not None:
                        count += 1
        return count

    def check_for_duplicate_links(self):
        for link in self.company_urls:
            while self.company_urls.count(link) > 1:
                self.company_urls.remove(link)
                print()
                warnprint('Há uma cópia de link: %s' % link)

    def get_links_from_workbook(self):
        links_sheet = self.workbook['Empresas']

        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value

        while link is not None:
            self.company_urls.append(link)

            line += 1
            link = links_sheet['C%i' % line].value

        if len(self.company_urls) == 0:
            print()
            checkprint('Todos os links do Excel já passaram pelo scraping!\nCaso queira recarregá-los, desative a configuração de "Apenas obter dados dos links cujos campos da linha estão vazios" e salve o arquivo\n')
            return not None
        else:
            return None

    def apply_links_sheet_style(self):
        self.apply_style_to_workbook_sheet(
            sheet=self.workbook['Empresas'], 
            verification_column='C', 
            starting_line=LINKS_TABLE_STARTING_LINE, 
            columns="BCDE"
        )
        self.apply_style_to_workbook_sheet(
            sheet=self.workbook['Empresas'], 
            alignment=CENTER_CELL_ALIGNMENT,
            font=BIG_FONT_CELL, 
            verification_column='C', 
            starting_line=LINKS_TABLE_STARTING_LINE, 
            columns="B"
        )

    def apply_style_to_workbook_sheet(self, sheet, verification_column, starting_line, columns, alignment=LEFT_CELL_ALIGNMENT, border=CELL_BORDER, font=NORMAL_FONT_CELL):
        line = starting_line
        link = sheet['%s%i' % (verification_column, line)].value
        while link is not None:
            for column in columns:
                cell = sheet['%s%i' % (column, line)]
                cell.alignment = alignment
                cell.border = border
                cell.font = font
            line += 1
            link = sheet['%s%i' % (verification_column, line)].value
        self.workbook.save(self.companies_excel_path)

    def find_company_by_id(self, company_id):
        for company in self.employees_json_data['empresas']:
            if company['company_id'] == company_id:
                return company

    def refresh_workbook_company_data(self, company_id):
        company = self.find_company_by_id(company_id)

        links_sheet = self.workbook['Empresas']

        line = LINKS_TABLE_STARTING_LINE
        link = links_sheet['C%i' % line].value
        
        while link is not None:

            if link == company['url']:

                links_sheet['B%i' % line] = '%i/%i' % (company['urls_de_funcionarios_obtidos'], company['quantidade_funcionarios'])
                links_sheet['D%i' % line] = company['nome']
                links_sheet['E%i' % line] = self.employees_json_path

                self.workbook.save(self.companies_excel_path)

            line += 1
            link = links_sheet['C%i' % line].value

    def login(self, response):
        return Http.FormRequest.from_response(
            response,
            formdata={
                'session_key': self.username,
                'session_password': self.password,
            },
            cookies=self.chrome_cookies,
            callback=self.check_login_response
        )

    def check_login_response(self, response):
        logged_in = False

        def loginerrorprint(x): return warnprint('Login falhou. %s Acesse o linkedin com essa conta para mais detalhes.\n' % x)

        print()

        if "Your account has been restricted" in str(response.body):
            loginerrorprint('Conta bloqueada pelo Linkedin por muitas tentativas.')
        elif "Let&#39;s do a quick security check" in str(response.body):
            loginerrorprint("Conta pede uma verificação de se é um robô.")
        elif "The login attempt seems suspicious." in str(response.body):
            loginerrorprint("Conta pede que seja copiado um código do email.")
        elif "that&#39;s not the right password" in str(response.body):
            loginerrorprint("A senha está errada.\nVerifique se o usuário e senha estão corretos.")
        elif "We’re unable to reach you" in str(response.body):
            loginerrorprint('O Linkedin pediu uma verificação de email.')
        else:
            logged_in = True
            checkprint("Login realizado. Vamos começar o crawling!\n")

        if logged_in:
            return self.initialized()
        else:
            return

    def start_requests(self):
        self._postinit_reqs = self.start_url_requests()
        return iterate_spider_output(self.init_request())

    def start_url_requests(self):
        for url in self.company_urls:
            yield self.cookie_request(
                url=url,
                callback=self.store_requests_preserving_priority
            )

    def get_company_included_array(self, response):
        body = str(response.body.decode('utf8'))

        birthIndex = body.rindex('&quot;videosTabVisible&quot;:')
        start = body[:birthIndex].rindex('<code ')
        end = body[start:].index('</code>') + start

        while (not body[start:end].startswith('{')) and start < end:
            start += 1

        while (not body[start:end].endswith('}')) and start < end:
            end -= 1

        if start >= end:
            whiteprint(
                'ERRO em get_company_included_array: não foi possivel obter dados do usuário em %s' % response.url)
            return None

        # save_to_file(
        #     response.url.split('/')[4] + '.html',
        #     convert_unicode(body[start:end], unicode_dict)
        # )

        return parse_text_to_json(body[start:end], unicode_dict, 'aa.json')['included']

    def get_employees_search_elements(self, response):

        body = str(response.body.decode('utf8'))

        birthIndex = body.rindex('&quot;com.linkedin.voyager.search.BlendedSearchCluster&quot;')
        start = body[:birthIndex].rindex('<code ')
        end = body[start:].index('</code>') + start

        while (not body[start:end].startswith('{')) and start < end:
            start += 1

        while (not body[start:end].endswith('}')) and start < end:
            end -= 1

        if start >= end:
            whiteprint(
                'ERRO em get_employees_data: não foi possivel obter dados do usuário em %s' % response.url)
            return None

        # save_to_file(
        #     'employee_search.json',
        #     convert_unicode(body[start:end], unicode_dict)
        # )

        blendedSearchClusterCollection = parse_text_to_json(body[start:end], unicode_dict, 'aa.json')['data']['elements']

        for blendedSearchCluster in blendedSearchClusterCollection:
            if blendedSearchCluster['type'] == 'SEARCH_HITS':
                return blendedSearchCluster['elements']

        return None

    def get_object_by_type(self, included_array, obj_type):
        array = []
        for obj in included_array:
            if obj['$type'] == obj_type:
                array.append(obj)
        return array

    def compare_employees(self, employee1, employee2):
        if ('url' in employee1) and ('url' in employee2):
            return employee1['url'] == employee2['url']
        else:
            return (employee1['localizacao_atual'] == employee2['localizacao_atual']) \
                and (employee1['cargo_atual'] == employee2['cargo_atual'])

    def should_employee_replace(self, old_employee, new_employee):
        if ('url' in old_employee) and ('url' in new_employee) and (old_employee['url'] == new_employee['url']):
            return False
        if 'url' in new_employee:
            return True
        return False

    def insert_employee_if_necessary(self, company_id, new_employee):
        for company in self.employees_json_data['empresas']:
            if company['company_id'] == company_id:

                for existing_employee in company['funcionarios']:

                    sameEmployees = self.compare_employees(existing_employee, new_employee)
                    shouldReplace = self.should_employee_replace(existing_employee, new_employee)

                    if sameEmployees:
                        if shouldReplace:
                            existing_employee.update(new_employee)
                        return

                company['funcionarios'].append(new_employee)
                return

    # Isso pode ser ativado quando a url não começa com www:
    def check_response_status(self, response):
        if response.status == 999:
            print()
            errorprint('Status 999. O Linkedin começou a restringir pedidos.\nO crawler será encerrado automaticamente.\n')
            raise CloseSpider('Spider encerrado manualmente')

    def store_requests_preserving_priority(self, response):
        self.stored_requests.extend(self.parse_company(response))
        self.companies_parsed += 1
        if self.companies_parsed == len(self.company_urls):
            self.stored_requests.sort(key=lambda elem: -elem.priority)
            for request in self.stored_requests:
                yield request

    def parse_company(self, response):
        self.check_response_status(response)
        
        company_included = self.get_company_included_array(response)

        company_info = self.get_object_by_type(
            company_included, 'com.linkedin.voyager.organization.Company')[0]

        company_id = company_info['companyEmployeesSearchPageUrl'].split('=')[-1]
        company_found = False
        for company_index in range(len(self.employees_json_data['empresas'])):
            if self.employees_json_data['empresas'][company_index]['company_id'] == company_id:
                company_found = True

                self.employees_json_data['empresas'][company_index].update({
                        'quantidade_funcionarios': company_info['staffCount'],
                    })

                break
                    

        if not company_found:
            self.employees_json_data['empresas'].append({
                'url': str(response.url),
                'dados_obtidos': False,
                'company_id': company_id,
                'nome': company_info['name'],
                'quantidade_funcionarios': company_info['staffCount'],
                'urls_de_funcionarios_obtidos': 0,
                'funcionarios': [],
            })

        self.refresh_workbook_company_data(company_id)

        employeeSearchPages = ceil(company_info['staffCount'] / 10)

        for page in range(employeeSearchPages):
            yield self.cookie_request(
                url='https://www.linkedin.com/search/results/people/?facetCurrentCompany=%%5B"%s"%%5D&page=%i' % (company_id, page + 1),
                callback=self.parse_employees_search,
                priority=(-page)
            )

    employees_search_parsed = 0

    def parse_employees_search(self, response):
        self.employees_search_parsed += 1

        if self.employees_search_parsed > 10:
            print()
            warnprint('Esta é uma versão de testes. Por segurança, aperte CRTL + C para encerrar, pois já foram acessadas 10 páginas de busca de funcionários.\n')
            raise CloseSpider('Spider encerrado manualmente')

        self.check_response_status(response)

        company_id = str(response.url).split('facetCurrentCompany=')[-1].split('&page')[0][6:-6]

        # save_to_file(
        #     response.url.split('/')[4] + '.html',
        #     response.body
        # )

        search_response = self.get_employees_search_elements(response)

        if search_response is not None:
            for company in self.employees_json_data['empresas']:
                if company['company_id'] == company_id:
                    urls_obtained = 0
                    for employee_data in search_response:

                        employee_url = employee_data['navigationUrl']

                        user_data = {
                            'localizacao_atual': employee_data['subline']['text'] if 'subline' in employee_data else None,
                            'cargo_atual': employee_data['headline']['text'] if 'headline' in employee_data else None,
                            'dados_obtidos': False
                        }

                        if '/in/UNKNOWN' not in employee_url:
                            urls_obtained += 1

                            user_data.update({
                                'url': employee_url
                            })

                        self.insert_employee_if_necessary(company_id, user_data)

                    company['urls_de_funcionarios_obtidos'] = self.count_employees_with_url(company_id)

                    conditionalprint = whiteprint if urls_obtained == 0 else checkprint

                    print()
                    conditionalprint(
                        '%i/%i URLs obtidos na página %s/%i de funcionários da empresa %s\n' 
                        % (
                            urls_obtained, 
                            len(search_response), 
                            response.url.split('&page=')[-1], 
                            ceil(company['quantidade_funcionarios'] / 10), 
                            company['nome']
                        )
                    )

                    save_to_file(
                        self.employees_json_path,
                        json.dumps(self.employees_json_data, indent=4),
                        dont_print=True
                    )

                    self.refresh_workbook_company_data(company_id)

                    return

            self.refresh_workbook_company_data(company_id)

            print()
            errorprint('Empresa não foi encontrada em %s\n' % self.employees_json_path)

        else:
            print()
            errorprint('Erro no parsing de lista de funcionários.\n')



def parse_text_to_json(text, replacements, filename):
    try:
        text = convert_unicode(text, replacements)
        return json.loads(text)
    except Exception:
        # save_to_file(
        #     filename,
        #     text
        # )
        return None


def convert_unicode(text, replacements):
    try:
        text = str(text)
        for unicode_char in list(replacements.keys()):
            for type in list(replacements[unicode_char].keys()):
                for element in replacements[unicode_char][type]:
                    text = text.replace(str(element), str(unicode_char))
    except Exception:
        print()
        errorprint(
            'convert_unicode: não foi possível converter os caracteres unicode.\n')
    return text


def read_json_file(path):
    try:
        f = open(path, "r+")
        data = json.loads(f.read())
        f.close()
        return data
    except json.decoder.JSONDecodeError:
        return None


def save_to_file(filename, element, dont_print=False):
    # element = str(element).replace("'", '"').replace('"s ', "'s ").replace(
    #     'True', 'true').replace('False', 'false').replace('None', 'null')
    with open(filename, 'wb') as f:
        f.write(str.encode(str(element)))
    if not dont_print: whiteprint('\n💽 Texto salvo como %s\n' % filename)
